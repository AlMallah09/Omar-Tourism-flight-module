import io
from datetime import date

from sqlalchemy import func
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.bookings.models import Booking
from app.users.models import User
from app.flights.models import Flight
from app.core.constants import (
    BookingStatus,
    PaymentStatus,
    FlightStatus
)


def to_excel_datetime(value):
    if value is None:
        return None

    if getattr(value, "tzinfo", None) is not None:
        return value.replace(tzinfo=None)

    return value


def apply_report_date_filter(
    query,
    column,
    start_date: date | None = None,
    end_date: date | None = None
):
    if start_date:
        query = query.filter(func.date(column) >= start_date)

    if end_date:
        query = query.filter(func.date(column) <= end_date)

    return query


def style_header_row(worksheet, row_number=1):
    for cell in worksheet[row_number]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


def auto_fit_worksheet(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        worksheet.column_dimensions[column_letter].width = min(
            max_length + 15,
            150
        )


def save_workbook_to_memory(workbook):
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def generate_bookings_excel(
    db: Session,
    start_date: date | None = None,
    end_date: date | None = None
):
    query = (
        db.query(Booking, User, Flight)
        .join(User, Booking.user_id == User.user_id)
        .join(Flight, Booking.flight_id == Flight.flight_id)
    )

    query = apply_report_date_filter(
        query,
        Booking.booking_date,
        start_date,
        end_date
    )

    bookings = (
        query
        .order_by(Booking.booking_id)
        .all()
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Bookings"

    worksheet.append([
        "Booking ID",
        "Customer Email",
        "Flight ID",
        "Airline",
        "Origin",
        "Destination",
        "Booking Status",
        "Payment Status",
        "Total Price"
    ])

    style_header_row(worksheet)

    for booking, user, flight in bookings:
        worksheet.append([
            booking.booking_id,
            user.email,
            flight.flight_id,
            flight.airline,
            flight.origin,
            flight.destination,
            booking.booking_status,
            booking.payment_status,
            booking.total_price
        ])

    auto_fit_worksheet(worksheet)

    return save_workbook_to_memory(workbook)


def generate_flights_excel(db: Session):
    flights = (
        db.query(Flight)
        .order_by(Flight.flight_id)
        .all()
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Flights"

    worksheet.append([
        "Flight ID",
        "Airline",
        "Origin",
        "Destination",
        "Departure Time",
        "Arrival Time",
        "Price",
        "Total Seats",
        "Seats Available",
        "Status"
    ])

    style_header_row(worksheet)

    for flight in flights:
        worksheet.append([
            flight.flight_id,
            flight.airline,
            flight.origin,
            flight.destination,
            to_excel_datetime(flight.departure_time),
            to_excel_datetime(flight.arrival_time),
            flight.price,
            flight.total_seats,
            flight.seats_available,
            flight.status
        ])

    auto_fit_worksheet(worksheet)

    return save_workbook_to_memory(workbook)


def generate_customers_excel(db: Session):
    customers = (
        db.query(
            User.user_id,
            User.email,
            func.count(Booking.booking_id).label("total_bookings"),
            func.coalesce(
                func.sum(Booking.total_price),
                0
            ).label("total_spent")
        )
        .outerjoin(
            Booking,
            Booking.user_id == User.user_id
        )
        .filter(User.role == "user")
        .group_by(
            User.user_id,
            User.email
        )
        .order_by(User.user_id)
        .all()
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Customers"

    worksheet.append([
        "Customer ID",
        "Email",
        "Total Bookings",
        "Total Booking Value"
    ])

    style_header_row(worksheet)

    for customer in customers:
        worksheet.append([
            customer.user_id,
            customer.email,
            customer.total_bookings,
            float(customer.total_spent)
        ])

    auto_fit_worksheet(worksheet)

    return save_workbook_to_memory(workbook)


def generate_revenue_excel(db: Session):
    total_booking_value = db.query(
        func.coalesce(
            func.sum(Booking.total_price),
            0
        )
    ).scalar()

    paid_revenue = (
        db.query(
            func.coalesce(
                func.sum(Booking.total_price),
                0
            )
        )
        .filter(
            Booking.payment_status == PaymentStatus.PAID.value
        )
        .scalar()
    )

    refunded_amount = (
        db.query(
            func.coalesce(
                func.sum(Booking.total_price),
                0
            )
        )
        .filter(
            Booking.payment_status == PaymentStatus.REFUNDED.value
        )
        .scalar()
    )

    total_bookings = db.query(Booking).count()

    average_booking_value = db.query(
        func.coalesce(
            func.avg(Booking.total_price),
            0
        )
    ).scalar()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Revenue"

    worksheet.append([
        "Metric",
        "Value"
    ])

    style_header_row(worksheet)

    metrics = [
        [
            "Total Booking Value",
            float(total_booking_value)
        ],
        [
            "Paid Revenue",
            float(paid_revenue)
        ],
        [
            "Refunded Amount",
            float(refunded_amount)
        ],
        [
            "Total Bookings",
            total_bookings
        ],
        [
            "Average Booking Value",
            round(float(average_booking_value), 2)
        ]
    ]

    for row in metrics:
        worksheet.append(row)

    auto_fit_worksheet(worksheet)

    return save_workbook_to_memory(workbook)


def generate_business_report_excel(
    db: Session,
    start_date: date | None = None,
    end_date: date | None = None
):
    workbook = Workbook()

    if start_date and end_date:
        reporting_period = f"{start_date} to {end_date}"
    elif start_date:
        reporting_period = f"From {start_date}"
    elif end_date:
        reporting_period = f"Up to {end_date}"
    else:
        reporting_period = "All Time"

    booking_stats_query = apply_report_date_filter(
        db.query(Booking),
        Booking.booking_date,
        start_date,
        end_date
    )

    total_bookings = booking_stats_query.count()

    confirmed_bookings = (
        booking_stats_query
        .filter(
            Booking.booking_status == BookingStatus.CONFIRMED.value
        )
        .count()
    )

    cancelled_bookings = (
        booking_stats_query
        .filter(
            Booking.booking_status == BookingStatus.CANCELLED.value
        )
        .count()
    )

    cancellation_rate = (
        (cancelled_bookings / total_bookings) * 100
        if total_bookings
        else 0
    )

    total_revenue = apply_report_date_filter(
        db.query(
            func.coalesce(
                func.sum(Booking.total_price),
                0
            )
        ),
        Booking.booking_date,
        start_date,
        end_date
    ).scalar()

    paid_revenue_query = db.query(
        func.coalesce(
            func.sum(Booking.total_price),
            0
        )
    ).filter(
        Booking.payment_status == PaymentStatus.PAID.value
    )

    paid_revenue = apply_report_date_filter(
        paid_revenue_query,
        Booking.booking_date,
        start_date,
        end_date
    ).scalar()

    refunded_amount_query = db.query(
        func.coalesce(
            func.sum(Booking.total_price),
            0
        )
    ).filter(
        Booking.payment_status == PaymentStatus.REFUNDED.value
    )

    refunded_amount = apply_report_date_filter(
        refunded_amount_query,
        Booking.booking_date,
        start_date,
        end_date
    ).scalar()

    average_booking_value = apply_report_date_filter(
        db.query(
            func.coalesce(
                func.avg(Booking.total_price),
                0
            )
        ),
        Booking.booking_date,
        start_date,
        end_date
    ).scalar()

    active_customers_query = (
        db.query(
            func.count(
                func.distinct(Booking.user_id)
            )
        )
        .join(
            User,
            Booking.user_id == User.user_id
        )
        .filter(User.role == "user")
    )

    active_customers = (
        apply_report_date_filter(
            active_customers_query,
            Booking.booking_date,
            start_date,
            end_date
        ).scalar()
        or 0
    )

    flights = (
        apply_report_date_filter(
            db.query(Flight),
            Flight.departure_time,
            start_date,
            end_date
        )
        .order_by(Flight.flight_id)
        .all()
    )

    total_flights = len(flights)

    active_flights = sum(
        1
        for flight in flights
        if flight.status == FlightStatus.SCHEDULED.value
    )

    total_capacity = sum(
        flight.total_seats or 0
        for flight in flights
    )

    total_available = sum(
        flight.seats_available or 0
        for flight in flights
    )

    booked_seats = total_capacity - total_available

    occupancy_rate = (
        (booked_seats / total_capacity) * 100
        if total_capacity
        else 0
    )

    summary_sheet = workbook.active
    summary_sheet.title = "Executive Summary"

    summary_sheet.append([
        "Omar Tourism Business Report"
    ])

    summary_sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    summary_sheet.append([
        "Reporting Period",
        reporting_period
    ])

    summary_sheet.append([])

    summary_sheet.append([
        "Metric",
        "Value"
    ])

    style_header_row(
        summary_sheet,
        row_number=4
    )

    summary_data = [
        ["Total Bookings", total_bookings],
        ["Confirmed Bookings", confirmed_bookings],
        ["Cancelled Bookings", cancelled_bookings],
        ["Cancellation Rate (%)", round(cancellation_rate, 2)],
        ["Total Booking Value", float(total_revenue)],
        ["Paid Revenue", float(paid_revenue)],
        ["Refunded Amount", float(refunded_amount)],
        [
            "Average Booking Value",
            round(float(average_booking_value), 2)
        ],
        ["Active Customers", active_customers],
        ["Flights in Reporting Period", total_flights],
        ["Scheduled Flights", active_flights],
        ["Total Seat Capacity", total_capacity],
        ["Booked Seats", booked_seats],
        ["Overall Occupancy Rate (%)", round(occupancy_rate, 2)]
    ]

    for row in summary_data:
        summary_sheet.append(row)

    bookings_sheet = workbook.create_sheet("Bookings")

    bookings_sheet.append([
        "Booking ID",
        "Customer Email",
        "Flight ID",
        "Airline",
        "Origin",
        "Destination",
        "Booking Status",
        "Payment Status",
        "Total Price",
        "Booking Date"
    ])

    style_header_row(bookings_sheet)

    booking_query = (
        db.query(Booking, User, Flight)
        .join(User, Booking.user_id == User.user_id)
        .join(Flight, Booking.flight_id == Flight.flight_id)
    )

    booking_query = apply_report_date_filter(
        booking_query,
        Booking.booking_date,
        start_date,
        end_date
    )

    bookings = (
        booking_query
        .order_by(Booking.booking_id)
        .all()
    )

    for booking, user, flight in bookings:
        bookings_sheet.append([
            booking.booking_id,
            user.email,
            flight.flight_id,
            flight.airline,
            flight.origin,
            flight.destination,
            booking.booking_status,
            booking.payment_status,
            booking.total_price,
            to_excel_datetime(booking.booking_date)
        ])

    flights_sheet = workbook.create_sheet("Flights")

    flights_sheet.append([
        "Flight ID",
        "Airline",
        "Origin",
        "Destination",
        "Departure Time",
        "Arrival Time",
        "Price",
        "Total Seats",
        "Seats Available",
        "Booked Seats",
        "Occupancy Rate (%)",
        "Status"
    ])

    style_header_row(flights_sheet)

    for flight in flights:
        flight_booked_seats = (
            flight.total_seats - flight.seats_available
        )

        flight_occupancy = (
            (flight_booked_seats / flight.total_seats) * 100
            if flight.total_seats
            else 0
        )

        flights_sheet.append([
            flight.flight_id,
            flight.airline,
            flight.origin,
            flight.destination,
            to_excel_datetime(flight.departure_time),
            to_excel_datetime(flight.arrival_time),
            flight.price,
            flight.total_seats,
            flight.seats_available,
            flight_booked_seats,
            round(flight_occupancy, 2),
            flight.status
        ])

    customers_sheet = workbook.create_sheet("Customers")

    customers_sheet.append([
        "Customer ID",
        "Email",
        "Total Bookings",
        "Total Booking Value"
    ])

    style_header_row(customers_sheet)

    customers_query = (
        db.query(
            User.user_id,
            User.email,
            func.count(
                Booking.booking_id
            ).label("total_bookings"),
            func.coalesce(
                func.sum(Booking.total_price),
                0
            ).label("total_spent")
        )
        .join(
            Booking,
            Booking.user_id == User.user_id
        )
        .filter(User.role == "user")
    )

    customers_query = apply_report_date_filter(
        customers_query,
        Booking.booking_date,
        start_date,
        end_date
    )

    customers = (
        customers_query
        .group_by(
            User.user_id,
            User.email
        )
        .order_by(User.user_id)
        .all()
    )

    for customer in customers:
        customers_sheet.append([
            customer.user_id,
            customer.email,
            customer.total_bookings,
            float(customer.total_spent)
        ])

    revenue_sheet = workbook.create_sheet("Revenue")

    revenue_sheet.append([
        "Metric",
        "Value"
    ])

    style_header_row(revenue_sheet)

    revenue_data = [
        ["Total Booking Value", float(total_revenue)],
        ["Paid Revenue", float(paid_revenue)],
        ["Refunded Amount", float(refunded_amount)],
        [
            "Average Booking Value",
            round(float(average_booking_value), 2)
        ]
    ]

    for row in revenue_data:
        revenue_sheet.append(row)

    analytics_sheet = workbook.create_sheet(
        "Business Analytics"
    )

    analytics_sheet.append([
        "Business Metric",
        "Value"
    ])

    style_header_row(analytics_sheet)

    average_bookings_per_customer = (
        total_bookings / active_customers
        if active_customers
        else 0
    )

    analytics_data = [
        ["Active Customers", active_customers],
        [
            "Average Bookings Per Active Customer",
            round(average_bookings_per_customer, 2)
        ],
        ["Total Bookings", total_bookings],
        ["Cancellation Rate (%)", round(cancellation_rate, 2)],
        ["Flights in Reporting Period", total_flights],
        ["Overall Flight Occupancy (%)", round(occupancy_rate, 2)]
    ]

    for row in analytics_data:
        analytics_sheet.append(row)

    for worksheet in workbook.worksheets:
        auto_fit_worksheet(worksheet)

    return save_workbook_to_memory(workbook)